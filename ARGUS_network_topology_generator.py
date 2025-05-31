"""
K@nj@k coding mood
"""

# ============================================================================
# IMPORTS ET CONFIGURATION INITIALE
# ============================================================================

# Importation des modules nécessaires pour le fonctionnement du script
import subprocess  # Pour exécuter des commandes système (ping)
from netmiko import ConnectHandler, NetmikoAuthenticationException, NetmikoTimeoutException  # Pour la connexion SSH aux équipements réseau
import platform  # Pour détecter le système d'exploitation (Windows/Linux)
from tabulate import tabulate  # Pour afficher les tableaux de manière lisible dans le terminal
from collections import defaultdict  # Pour créer des dictionnaires avec des valeurs par défaut
from openpyxl import Workbook  # Pour créer et manipuler des fichiers Excel
from datetime import datetime  # Pour générer des timestamps
import re  # Pour les expressions régulières
import graphviz  # Pour générer la représentation graphique du réseau
import json  # Pour l'export JSON

# Importation de la configuration
from config import (
    DEVICES,              # Liste des équipements à interroger
    DEFAULT_CREDENTIALS,  # Credentials par défaut
    TIMEOUT,             # Délai d'attente pour la connexion
    SESSION_TIMEOUT,     # Délai d'attente pour la session
    DETECTION_COMMANDS   # Commandes pour détecter le type d'équipement
)

# ============================================================================
# FONCTIONS DE VÉRIFICATION ET DE CONNEXION
# ============================================================================

def is_device_up(ip):
    """
    Vérifie si un équipement réseau est accessible via ping.
    
    Args:
        ip (str): L'adresse IP de l'équipement à tester
        
    Returns:
        bool: True si l'équipement répond au ping, False sinon
    """
    try:
        # Détermine le paramètre à utiliser pour la commande ping selon le système d'exploitation
        # -n pour Windows (nombre de paquets)
        # -c pour Linux (nombre de paquets)
        param = '-n' if platform.system().lower() == 'windows' else '-c'
        
        # Exécute la commande ping avec un seul paquet
        # stdout=subprocess.DEVNULL : masque la sortie de la commande
        result = subprocess.run(["ping", param, "1", ip], stdout=subprocess.DEVNULL)
        
        # Retourne True si le code de retour est 0 (ping réussi)
        # Retourne False dans tous les autres cas
        return result.returncode == 0
    except Exception:
        # En cas d'erreur quelconque, considère que l'équipement n'est pas accessible
        return False

def detect_device_type(conn):
    """
    Détecte automatiquement le type d'équipement en essayant différentes commandes.
    
    Args:
        conn: Connexion SSH établie
        
    Returns:
        dict: Configuration de l'équipement détecté ou None si non détecté
    """
    for device_config in DETECTION_COMMANDS:
        try:
            # Essaie d'exécuter la commande de détection
            output = conn.send_command(
                device_config["command"],
                expect_string=r"#",
                read_timeout=TIMEOUT
            )
            
            # Vérifie si la sortie correspond au modèle attendu
            if re.search(device_config["model_pattern"], output):
                return device_config
        except:
            continue
    
    return None

def get_device_info(ip):
    """
    Récupère les informations d'un équipement réseau via SSH.
    Cette fonction :
    1. Établit une connexion SSH en tant que Cisco IOS
    2. Vérifie que LLDP est activé
    3. Récupère le hostname et les informations sur les voisins LLDP
    
    Args:
        ip (str): L'adresse IP de l'équipement à interroger
        
    Returns:
        tuple: (hostname, neighbors, device_info) où :
            - hostname est le nom de l'équipement
            - neighbors est une liste des voisins [nom, interface_locale, interface_distante]
            - device_info est un dict contenant les infos de l'équipement (type, vendor, model)
            Retourne (None, None, None) en cas d'erreur
    """
    try:
        print(f"\n[INFO] Tentative de connexion à {ip} en tant que cisco_ios...")
        # Configuration de base de la connexion SSH
        device = {
            'device_type': 'cisco_ios',
            'host': ip,
            'username': DEFAULT_CREDENTIALS["username"],
            'password': DEFAULT_CREDENTIALS["password"],
            'secret': DEFAULT_CREDENTIALS['secret'],
            'timeout': TIMEOUT,
            'session_timeout': SESSION_TIMEOUT
        }
        
        # Tentative de connexion
        conn = ConnectHandler(**device)
        print(f"[OK] Connexion SSH réussie à {ip}")
        
        # Passage en mode privilégié
        conn.enable()
        
        # Détection du type d'équipement
        device_config = detect_device_type(conn)
        if not device_config:
            print(f"[ERREUR] Détection du type d'équipement échouée pour {ip}")
            conn.disconnect()
            return None, None, None
        
        # Vérification que LLDP est activé
        print(f"[DEBUG] Vérification du statut LLDP sur {ip}...")
        lldp_status = conn.send_command("show lldp", expect_string=r"#", read_timeout=TIMEOUT)
        print(f"[DEBUG] Sortie de 'show lldp':\n{lldp_status}\n{'='*60}")
        
        if "LLDP is not enabled" in lldp_status:
            print(f"[ERREUR] LLDP n'est pas activé sur {ip}")
            conn.disconnect()
            return None, None, None
        
        # Récupérer le hostname
        hostname_output = conn.send_command("show running-config | include hostname", expect_string=r"#", read_timeout=20)
        hostname = "Unknown"
        for line in hostname_output.splitlines():
            if "hostname" in line.lower():
                hostname = line.split("hostname")[1].strip()
                break
        print(f"[DEBUG] Hostname détecté: {hostname}")
        
        # Récupération des voisins LLDP
        try:
            # Essai d'abord avec la commande simple
            print(f"[DEBUG] Tentative avec 'show lldp neighbors' sur {ip}...")
            output = conn.send_command(
                device_config['neighbors_cmd'],
                expect_string=r"#",
                read_timeout=TIMEOUT
            )
            print(f"[DEBUG] Sortie brute de 'show lldp neighbors':\n{output}\n{'='*60}")
            
            # Si pas de voisins trouvés ou sortie vide, essayer la commande détaillée
            if not output.strip() or "No LLDP neighbors" in output:
                print(f"[DEBUG] Aucun voisin trouvé avec la commande simple, tentative avec 'show lldp neighbors detail'...")
                output = conn.send_command(
                    device_config['neighbors_alt_cmd'],
                    expect_string=r"#",
                    read_timeout=TIMEOUT
                )
                print(f"[DEBUG] Sortie brute de 'show lldp neighbors detail':\n{output}\n{'='*60}")
            
            if "No LLDP neighbors" in output:
                print(f"[INFO] Aucun voisin LLDP trouvé sur {ip}")
                conn.disconnect()
                return hostname, [], {
                    'device_type': device_config['device_type'],
                    'vendor': device_config['vendor'],
                    'model': 'Unknown',
                    'protocol': 'LLDP'
                }
                
        except Exception as e:
            print(f"[ERREUR] Récupération des voisins LLDP échouée pour {ip}: {e}")
            conn.disconnect()
            return hostname, [], {
                'device_type': device_config['device_type'],
                'vendor': device_config['vendor'],
                'model': 'Unknown',
                'protocol': 'LLDP'
            }
        
        # Parsing des voisins LLDP
        neighbors = []
        parse_config = device_config['neighbors_parse']
        
        print(f"[DEBUG] Début du parsing des voisins LLDP pour {ip}...")
        
        # Ignorer les lignes d'en-tête et de capacité
        lines = output.splitlines()
        start_parsing = False
        
        for line in lines:
            line = line.strip()
            
            # Ignorer les lignes vides
            if not line:
                continue
                
            # Ignorer les lignes d'en-tête et de capacité
            if any(x in line for x in ["Capability codes:", "Device ID", "Total entries"]):
                continue
                
            # Ignorer les lignes qui commencent par des parenthèses (capability codes)
            if line.startswith("("):
                continue
                
            # Si on trouve une ligne qui commence par un nom d'équipement, c'est un voisin
            parts = line.split()
            if len(parts) >= 4 and not parts[0].startswith("("):  # Une ligne de voisin doit avoir au moins 4 colonnes et ne pas commencer par (
                try:
                    neighbor_name = parts[0]
                    local_intf = parts[1]
                    remote_intf = parts[-1]  # La dernière colonne est le Port ID
                    
                    # Nettoyage du nom du voisin (suppression du domaine si présent)
                    neighbor_name = neighbor_name.split('.')[0]
                    
                    # Vérifier que ce n'est pas une ligne de capability
                    if not any(x in neighbor_name for x in ["Router", "Bridge", "Telephone", "WLAN", "Repeater", "Station"]):
                        neighbors.append([neighbor_name, local_intf, remote_intf])
                        print(f"[DEBUG] Voisin ajouté: {neighbor_name} ({local_intf} -> {remote_intf})")
                except Exception as e:
                    print(f"[DEBUG] Erreur lors du parsing de la ligne: {line}")
                    print(f"[DEBUG] Erreur: {e}")
                    continue
        
        print(f"[DEBUG] Nombre total de voisins trouvés pour {ip}: {len(neighbors)}")
        
        conn.disconnect()
        
        # Extraction du modèle depuis la sortie de la commande de détection
        model = 'Unknown'
        try:
            model_match = re.search(device_config['model_pattern'], output)
            if model_match:
                model = model_match.group(0)
        except Exception as e:
            print(f"[ERREUR] Extraction du modèle échouée pour {ip}: {e}")
            
        return hostname, neighbors, {
            'device_type': device_config['device_type'],
            'vendor': device_config['vendor'],
            'model': model,
            'protocol': 'LLDP'
        }
        
    except Exception as e:
        print(f"[ERREUR] Connexion SSH échouée à {ip}: {e}")
        return None, None, None

def extract_device_name(hostname):
    """
    Extrait le nom de l'équipement sans le domaine.
    
    Args:
        hostname (str): Le nom complet de l'équipement (ex: 'AN-SW-DIST-01.my-intranet.local')
        
    Returns:
        str: Le nom de l'équipement sans le domaine
    """
    return hostname.split('.')[0] if hostname else hostname

def create_network_map():
    """
    Fonction principale qui crée la cartographie du réseau.
    Cette fonction :
    1. Vérifie la disponibilité de chaque équipement
    2. Détecte automatiquement le type de chaque équipement
    3. Récupère les informations de chaque équipement accessible
    4. Génère un tableau de la topologie
    5. Affiche un résumé des connexions
    6. Exporte les données dans un fichier Excel
    """
    network_map = defaultdict(list)
    
    # Parcours de tous les équipements de la liste
    for ip in DEVICES:
        print(f"\n[TEST] Ping de {ip}...")
        if is_device_up(ip):
            print(f"[OK] {ip} répond au ping.")
            hostname, neighbors, device_info = get_device_info(ip)
            if hostname and device_info:
                # Extraction du nom court de l'équipement
                short_hostname = extract_device_name(hostname)
                network_map[short_hostname] = {
                    'ip': ip,
                    'device_type': device_info['device_type'],
                    'vendor': device_info['vendor'],
                    'model': device_info['model'],
                    'protocol': device_info['protocol'],
                    'neighbors': [(extract_device_name(n[0]), n[1], n[2]) for n in neighbors]
                }
            else:
                print(f"[INFO] Aucun voisin ou information récupérée pour {ip}.")
        else:
            print(f"[ERREUR] {ip} ne répond pas au ping.")
    
    # Création du tableau de la topologie (version non commentée)
    topology = []
    for device, info in network_map.items():
        print(f"DEBUG: {device} neighbors: {info['neighbors']}")
        for neighbor in info['neighbors']:
            topology.append([device, info['ip'], neighbor[0], neighbor[1], neighbor[2], info['protocol']])
    
    # Affichage des résultats si des informations ont été trouvées
    if topology:
        print("\nTopologie du réseau:")
        print(tabulate(topology, headers=["Équipement", "IP", "Voisin", "Interface Locale", "Interface Distante", "Protocole"], tablefmt="simple"))
        
        # Affichage du résumé des connexions (version non commentée)
        print("\nRésumé des connexions:")
        connections = defaultdict(set)
        for row in topology:
            device = row[0]
            neighbor = row[2]
            connections[device].add(neighbor)
        
        for device, neighbors in connections.items():
            print(f"{device} est connecté à: {', '.join(neighbors)}")
            
        # Création du fichier Excel (version non commentée)
        wb = Workbook()
        ws = wb.active
        ws.title = "Topologie"
        
        # Ajout des en-têtes
        headers = ["Équipement", "IP", "Voisin", "Interface Locale", "Interface Distante", "Protocole"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Ajout des données
        for row_idx, row_data in enumerate(topology, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustement de la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarde du fichier Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"topologies/topology_{timestamp}.xlsx"
        wb.save(filename)
        print(f"\nTopologie exportée dans le fichier : {filename}")
        
        # Export de la topologie en JSON
        json_data = {
            "timestamp": timestamp,
            "devices": {
                device: {
                    "ip": info["ip"],
                    "device_type": info["device_type"],
                    "vendor": info["vendor"],
                    "model": info["model"],
                    "protocol": info["protocol"],
                    "neighbors": [
                        {
                            "name": n[0],
                            "local_interface": n[1],
                            "remote_interface": n[2]
                        } for n in info["neighbors"]
                    ]
                } for device, info in network_map.items()
            }
        }
        
        json_filename = f"topologies/topology_{timestamp}.json"
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=4, ensure_ascii=False)
        print(f"Topologie exportée dans le fichier JSON : {json_filename}")
        
        # Création du graphique avec Graphviz
        dot = graphviz.Graph(comment='Topologie Réseau', engine='dot')
        
        # Configuration globale du graphe
        dot.attr(overlap='false', splines='true', rankdir='TB')
        
        # Style par défaut pour les nœuds
        dot.attr('node', shape='box', style='filled', fillcolor='lightblue',
                fontname='Helvetica', fontsize='10')
        
        # Style par défaut pour les arêtes
        dot.attr('edge', fontsize='8', fontname='Helvetica')

        # Identifier automatiquement les switches de distribution et d'accès
        distribution_switches = set()
        access_switches = set()
        connections_count = defaultdict(int)

        # Compter le nombre de connexions pour chaque équipement
        for row in topology:
            device = row[0]
            neighbor = row[2]
            connections_count[device] += 1
            connections_count[neighbor] += 1

        # Classifier les switches basé sur leur nombre de connexions
        for device, count in connections_count.items():
            if count > 1:  # Si un switch a plusieurs connexions, c'est probablement un switch de distribution
                distribution_switches.add(device)
            else:
                access_switches.add(device)

        # Créer des sous-graphes pour organiser la hiérarchie
        with dot.subgraph(name='cluster_0') as dist:
            dist.attr(rank='min')  # Force les switches de distribution en haut
            for switch in distribution_switches:
                dist.node(switch, switch)

        with dot.subgraph(name='cluster_1') as acc:
            acc.attr(rank='max')  # Force les switches d'accès en bas
            for switch in access_switches:
                acc.node(switch, switch)
        
        # Ajout des connexions avec les détails des interfaces
        seen_edges = set()  # Pour éviter les doublons
        for row in topology:
            device = row[0]
            neighbor = row[2]
            local_intf = row[3]
            remote_intf = row[4]
            
            # Création d'une clé unique pour chaque liaison
            edge_key = tuple(sorted([device, neighbor]))
            if edge_key not in seen_edges:
                # Création de deux labels, un pour chaque extrémité
                dot.edge(device, neighbor, headlabel=remote_intf, taillabel=local_intf, 
                        color='darkblue', fontsize='8', labeldistance='2.0', 
                        labelfloat='false', labelangle='45')
                seen_edges.add(edge_key)
        
        # Configuration supplémentaire pour améliorer le rendu
        dot.attr(dpi='300')  # Haute résolution
        
        # Sauvegarde du graphique
        graph_filename = f"topologies/topology_{timestamp}"
        dot.render(graph_filename, format='png', cleanup=True)
        print(f"Représentation graphique exportée dans le fichier : {graph_filename}.png")
    else:
        print("Aucune information de topologie trouvée.")

# Exécution de la cartographie du réseau
create_network_map()
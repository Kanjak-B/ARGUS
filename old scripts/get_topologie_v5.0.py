import subprocess
from netmiko import ConnectHandler, NetmikoAuthenticationException, NetmikoTimeoutException
import platform
from tabulate import tabulate
from collections import defaultdict
from openpyxl import Workbook
from datetime import datetime
import graphviz

# Liste des IP
devices = [
    "10.103.0.10",
    # "10.103.0.11",
    "10.103.0.12",
    "10.103.0.100",
    # "10.103.0.254"
]

# Credentials SSH
username = "admin"
password = "admin"
admin_password = "admin"  # pour le mode enable

def is_device_up(ip):
    try:
        param = '-n' if platform.system().lower() == 'windows' else '-c'
        result = subprocess.run(["ping", param, "1", ip], stdout=subprocess.DEVNULL)
        return result.returncode == 0
    except Exception:
        return False

def get_device_info(ip):
    device = {
        'device_type': 'cisco_ios',
        'host': ip,
        'username': username,
        'password': password,
        'secret': admin_password,
        'timeout': 20,
        'session_timeout': 20
    }
    
    try:
        conn = ConnectHandler(**device)
        conn.enable()
        
        # Récupérer le hostname
        hostname_output = conn.send_command("show running-config | include hostname", expect_string=r"#", read_timeout=20)
        hostname = "Unknown"
        for line in hostname_output.splitlines():
            if "hostname" in line.lower():
                hostname = line.split("hostname")[1].strip()
                break
        
        # Essayer CDP d'abord
        output = conn.send_command("show cdp neighbors detail", expect_string=r"#", read_timeout=20)
        protocol = "CDP"
        
        if "Invalid input" in output or "CDP is not enabled" in output:
            # Essayer LLDP si CDP ne fonctionne pas
            output = conn.send_command("show lldp neighbors detail", expect_string=r"#", read_timeout=20)
            protocol = "LLDP"
            if "Invalid input" in output or "LLDP is not enabled" in output:
                return None, None, None
        
        # Parser les voisins
        neighbors = []
        current_neighbor = None
        current_local_intf = None
        current_remote_intf = None
        
        for line in output.splitlines():
            line = line.strip()
            
            if protocol == "CDP":
                if "Device ID:" in line:
                    if current_neighbor and current_local_intf and current_remote_intf:
                        neighbors.append([current_neighbor, current_local_intf, current_remote_intf])
                    current_neighbor = line.split("Device ID:")[1].strip()
                    if '.' in current_neighbor:
                        current_neighbor = current_neighbor.split('.')[0]
                    if ' ' in current_neighbor:
                        current_neighbor = current_neighbor.split(' ')[0]
                elif "Interface:" in line and "Port ID (outgoing port):" in line:
                    # Les deux infos sont sur la même ligne
                    parts = line.split(',')
                    current_local_intf = parts[0].split("Interface:")[1].strip()
                    current_remote_intf = parts[1].split("Port ID (outgoing port):")[1].strip()
                    if current_neighbor and current_local_intf and current_remote_intf:
                        neighbors.append([current_neighbor, current_local_intf, current_remote_intf])
                        current_neighbor = None
                        current_local_intf = None
                        current_remote_intf = None
                elif "Interface:" in line:
                    current_local_intf = line.split("Interface:")[1].strip()
                elif "Port ID (outgoing port):" in line:
                    current_remote_intf = line.split("Port ID (outgoing port):")[1].strip()
                    if current_neighbor and current_local_intf and current_remote_intf:
                        neighbors.append([current_neighbor, current_local_intf, current_remote_intf])
                        current_neighbor = None
                        current_local_intf = None
                        current_remote_intf = None
            
            elif protocol == "LLDP":
                if "System Name:" in line:
                    if current_neighbor and current_local_intf and current_remote_intf:
                        neighbors.append([current_neighbor, current_local_intf, current_remote_intf])
                    current_neighbor = line.split("System Name:")[1].strip()
                elif "Local Interface:" in line:
                    current_local_intf = line.split("Local Interface:")[1].strip()
                elif "Remote Interface:" in line:
                    current_remote_intf = line.split("Remote Interface:")[1].strip()
                    if current_neighbor and current_local_intf:
                        neighbors.append([current_neighbor, current_local_intf, current_remote_intf])
                        current_neighbor = None
                        current_local_intf = None
                        current_remote_intf = None
        
        # Ajouter le dernier voisin s'il est complet
        if current_neighbor and current_local_intf and current_remote_intf:
            neighbors.append([current_neighbor, current_local_intf, current_remote_intf])
        
        conn.disconnect()
        return hostname, neighbors, protocol
        
    except Exception:
        return None, None, None

def create_network_map():
    # Dictionnaire pour stocker les informations de chaque équipement
    network_map = defaultdict(list)
    
    # Collecter les informations de tous les équipements
    for ip in devices:
        if is_device_up(ip):
            hostname, neighbors, protocol = get_device_info(ip)
            if hostname and neighbors:
                network_map[hostname] = {
                    'ip': ip,
                    'protocol': protocol,
                    'neighbors': neighbors
                }
    
    # Créer le tableau de la topologie
    topology = []
    for device, info in network_map.items():
        for neighbor in info['neighbors']:
            topology.append([
                device,
                info['ip'],
                neighbor[0],  # nom du voisin
                neighbor[1],  # interface locale
                neighbor[2],  # interface distante
                info['protocol']
            ])
    
    # Afficher la topologie
    if topology:
        print("\nTopologie du réseau:")
        print(tabulate(
            topology,
            headers=["Équipement", "IP", "Voisin", "Interface Locale", "Interface Distante", "Protocole"],
            tablefmt="simple"
        ))
        
        # Afficher un résumé des connexions
        print("\nRésumé des connexions:")
        connections = defaultdict(set)
        for row in topology:
            device = row[0]
            neighbor = row[2]
            connections[device].add(neighbor)
        
        for device, neighbors in connections.items():
            print(f"{device} est connecté à: {', '.join(neighbors)}")
            
        # Créer le fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Topologie"
        
        # Ajouter les en-têtes
        headers = ["Équipement", "IP", "Voisin", "Interface Locale", "Interface Distante", "Protocole"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Ajouter les données
        for row_idx, row_data in enumerate(topology, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder le fichier Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"excel/topology_v1_{timestamp}.xlsx"
        wb.save(excel_filename)
        print(f"\nTopologie exportée dans le fichier Excel : {excel_filename}")
        
        # Créer la représentation graphique
        dot = graphviz.Digraph(comment='Topologie Réseau')
        dot.attr(rankdir='LR')  # Direction gauche à droite
        
        # Style des nœuds
        dot.attr('node', shape='box', style='filled', fillcolor='lightblue')
        
        # Ajouter les nœuds (équipements)
        for device, info in network_map.items():
            label = f"{device}\n{info['ip']}"
            dot.node(device, label)
        
        # Ajouter les connexions
        for row in topology:
            device = row[0]
            neighbor = row[2]
            local_intf = row[3]
            remote_intf = row[4]
            label = f"{local_intf} - {remote_intf}"
            dot.edge(device, neighbor, label=label)
        
        # Sauvegarder le graphique
        graph_filename = f"excel/topology_v1_{timestamp}"
        dot.render(graph_filename, format='png', cleanup=True)
        print(f"Représentation graphique exportée dans le fichier : {graph_filename}.png")
        
    else:
        print("Aucune information de topologie trouvée.")

# Exécuter la cartographie
create_network_map() 
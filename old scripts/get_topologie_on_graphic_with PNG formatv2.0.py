"""
K@nj@k coding mood
"""

# ============================================================================
# IMPORTS ET CONFIGURATION INITIALE
# ============================================================================

# Importation des modules nécessaires pour le fonctionnement du script
import subprocess  # Pour exécuter des commandes système (ping)
from netmiko import ConnectHandler, NetmikoAuthenticationException, NetmikoTimeoutException  # Pour la connexion SSH aux équipements réseau
import platform  # Pour détecter le système d'exploitation (Windows/Linux)
from tabulate import tabulate  # Pour afficher les tableaux de manière lisible dans le terminal
from collections import defaultdict  # Pour créer des dictionnaires avec des valeurs par défaut
from openpyxl import Workbook  # Pour créer et manipuler des fichiers Excel
from datetime import datetime  # Pour générer des timestamps
import re  # Pour les expressions régulières
import graphviz  # Pour générer la représentation graphique du réseau

# Importation de la configuration
from config import (
    DEVICES,              # Liste des équipements à interroger
    DEFAULT_CREDENTIALS,  # Credentials par défaut
    TIMEOUT,             # Délai d'attente pour la connexion
    SESSION_TIMEOUT,     # Délai d'attente pour la session
    DETECTION_COMMANDS   # Commandes pour détecter le type d'équipement
)

# ============================================================================
# FONCTIONS DE VÉRIFICATION ET DE CONNEXION
# ============================================================================

def is_device_up(ip):
    """
    Vérifie si un équipement réseau est accessible via ping.
    
    Args:
        ip (str): L'adresse IP de l'équipement à tester
        
    Returns:
        bool: True si l'équipement répond au ping, False sinon
    """
    try:
        # Détermine le paramètre à utiliser pour la commande ping selon le système d'exploitation
        # -n pour Windows (nombre de paquets)
        # -c pour Linux (nombre de paquets)
        param = '-n' if platform.system().lower() == 'windows' else '-c'
        
        # Exécute la commande ping avec un seul paquet
        # stdout=subprocess.DEVNULL : masque la sortie de la commande
        result = subprocess.run(["ping", param, "1", ip], stdout=subprocess.DEVNULL)
        
        # Retourne True si le code de retour est 0 (ping réussi)
        # Retourne False dans tous les autres cas
        return result.returncode == 0
    except Exception:
        # En cas d'erreur quelconque, considère que l'équipement n'est pas accessible
        return False

def detect_device_type(conn):
    """
    Détecte automatiquement le type d'équipement en essayant différentes commandes.
    
    Args:
        conn: Connexion SSH établie
        
    Returns:
        dict: Configuration de l'équipement détecté ou None si non détecté
    """
    for device_config in DETECTION_COMMANDS:
        try:
            # Essaie d'exécuter la commande de détection
            output = conn.send_command(
                device_config["command"],
                expect_string=r"#",
                read_timeout=TIMEOUT
            )
            
            # Vérifie si la sortie correspond au modèle attendu
            if re.search(device_config["model_pattern"], output):
                return device_config
        except:
            continue
    
    return None

def get_device_info(ip):
    """
    Récupère les informations d'un équipement réseau via SSH.
    Cette fonction :
    1. Établit une connexion SSH en essayant chaque type d'équipement connu
    2. Détecte automatiquement le type d'équipement
    3. Récupère le hostname et les informations sur les voisins
    
    Args:
        ip (str): L'adresse IP de l'équipement à interroger
        
    Returns:
        tuple: (hostname, neighbors, device_info) où :
            - hostname est le nom de l'équipement
            - neighbors est une liste des voisins [nom, interface_locale, interface_distante]
            - device_info est un dict contenant les infos de l'équipement (type, vendor, model)
            Retourne (None, None, None) en cas d'erreur
    """
    # Liste des types d'équipements à essayer
    device_types = ["cisco_ios", "fortinet"]
    
    for device_type in device_types:
        try:
            print(f"[INFO] Tentative de connexion à {ip} en tant que {device_type}...")
            # Configuration de base de la connexion SSH
            device = {
                'device_type': device_type,
                'host': ip,
                'username': DEFAULT_CREDENTIALS["username"],
                'password': DEFAULT_CREDENTIALS["password"],
                'timeout': TIMEOUT,
                'session_timeout': SESSION_TIMEOUT
            }
            
            # Ajout du secret pour Cisco
            if device_type == "cisco_ios":
                device['secret'] = DEFAULT_CREDENTIALS['secret']
            
            # Tentative de connexion
            conn = ConnectHandler(**device)
            print(f"[OK] Connexion SSH réussie à {ip} ({device_type})")
            
            # Si c'est un équipement Cisco, passage en mode privilégié
            if device_type == "cisco_ios":
                conn.enable()
            
            # Détection du type d'équipement
            device_config = detect_device_type(conn)
            if not device_config:
                print(f"[ERREUR] Détection du type d'équipement échouée pour {ip} (type essayé: {device_type})")
                conn.disconnect()
                continue
            
            # Si le type détecté ne correspond pas au type essayé, on continue
            if device_config['device_type'] != device_type:
                print(f"[INFO] Type détecté ({device_config['device_type']}) différent du type essayé ({device_type}) pour {ip}")
                conn.disconnect()
                continue
            
            # Récupérer le hostname via la commande show running-config | include hostname (comme dans get_topologie_on_graphic_v2.0.py)
            hostname_output = conn.send_command("show running-config | include hostname", expect_string=r"#", read_timeout=20)
            hostname = "Unknown"
            for line in hostname_output.splitlines():
                if "hostname" in line.lower():
                    hostname = line.split("hostname")[1].strip()
                    break
            
            # Récupération des voisins
            try:
                output = conn.send_command(
                    device_config['neighbors_cmd'],
                    expect_string=r"#",
                    read_timeout=TIMEOUT
                )
                print(f"[DEBUG] Sortie brute de la commande de voisinage sur {ip} (cmd: {device_config['neighbors_cmd']}):\n{output}\n{'='*60}")
                protocol = device_config['neighbors_parse']['protocol']
                # Si le protocole principal ne fonctionne pas et qu'un protocole alternatif est défini
                if ("Invalid input" in output or "command not found" in output) and device_config['neighbors_alt_cmd']:
                    output = conn.send_command(
                        device_config['neighbors_alt_cmd'],
                        expect_string=r"#",
                        read_timeout=TIMEOUT
                    )
                    print(f"[DEBUG] Sortie brute de la commande de voisinage sur {ip} (cmd: {device_config['neighbors_alt_cmd']}):\n{output}\n{'='*60}")
                    protocol = device_config['neighbors_parse']['alt_protocol']
                    if "Invalid input" in output or "command not found" in output:
                        print(f"[ERREUR] Commandes de voisinage non supportées sur {ip}")
                        conn.disconnect()
                        return hostname, [], {
                            'device_type': device_config['device_type'],
                            'vendor': device_config['vendor'],
                            'model': 'Unknown',
                            'protocol': None
                        }
            except Exception as e:
                print(f"[ERREUR] Récupération des voisins échouée pour {ip}: {e}")
                conn.disconnect()
                return hostname, [], {
                    'device_type': device_config['device_type'],
                    'vendor': device_config['vendor'],
                    'model': 'Unknown',
                    'protocol': None
                }
            
            # Parsing des voisins
            neighbors = []
            parse_config = device_config['neighbors_parse']
            # On découpe la sortie en blocs de voisins
            neighbor_blocks = output.split(parse_config['device_id'])
            for block in neighbor_blocks[1:]:  # Le premier split est avant le premier voisin
                lines = block.splitlines()
                neighbor_name = lines[0].strip().split()[0] if lines else None
                local_intf = None
                remote_intf = None
                for line in lines:
                    if parse_config['local_intf'] in line:
                        # Exemple: 'Interface: Ethernet0/0,  Port ID (outgoing port): Ethernet3/3'
                        local_intf = line.split(parse_config['local_intf'])[1].split(',')[0].strip()
                        # Cherche aussi le Port ID sur la même ligne
                        if parse_config['remote_intf'] in line:
                            remote_intf = line.split(parse_config['remote_intf'])[1].strip()
                    elif parse_config['remote_intf'] in line:
                        remote_intf = line.split(parse_config['remote_intf'])[1].strip()
                if neighbor_name and local_intf and remote_intf:
                    neighbors.append([neighbor_name, local_intf, remote_intf])
            conn.disconnect()
            # Extraction du modèle depuis la sortie de la commande de détection
            model = 'Unknown'
            try:
                model_match = re.search(device_config['model_pattern'], output)
                if model_match:
                    model = model_match.group(0)
            except Exception as e:
                print(f"[ERREUR] Extraction du modèle échouée pour {ip}: {e}")
            return hostname, neighbors, {
                'device_type': device_config['device_type'],
                'vendor': device_config['vendor'],
                'model': model,
                'protocol': protocol
            }
        except Exception as e:
            print(f"[ERREUR] Connexion SSH échouée à {ip} en tant que {device_type}: {e}")
            continue
    print(f"Impossible de se connecter à {ip} avec les types d'équipements connus")
    return None, None, None

def extract_device_name(hostname):
    """
    Extrait le nom de l'équipement sans le domaine.
    
    Args:
        hostname (str): Le nom complet de l'équipement (ex: 'AN-SW-DIST-01.my-intranet.local')
        
    Returns:
        str: Le nom de l'équipement sans le domaine
    """
    return hostname.split('.')[0] if hostname else hostname

def create_network_map():
    """
    Fonction principale qui crée la cartographie du réseau.
    Cette fonction :
    1. Vérifie la disponibilité de chaque équipement
    2. Détecte automatiquement le type de chaque équipement
    3. Récupère les informations de chaque équipement accessible
    4. Génère un tableau de la topologie
    5. Affiche un résumé des connexions
    6. Exporte les données dans un fichier Excel
    """
    network_map = defaultdict(list)
    
    # Parcours de tous les équipements de la liste
    for ip in DEVICES:
        print(f"\n[TEST] Ping de {ip}...")
        if is_device_up(ip):
            print(f"[OK] {ip} répond au ping.")
            hostname, neighbors, device_info = get_device_info(ip)
            if hostname and device_info:
                # Extraction du nom court de l'équipement
                short_hostname = extract_device_name(hostname)
                network_map[short_hostname] = {
                    'ip': ip,
                    'device_type': device_info['device_type'],
                    'vendor': device_info['vendor'],
                    'model': device_info['model'],
                    'protocol': device_info['protocol'],
                    'neighbors': [(extract_device_name(n[0]), n[1], n[2]) for n in neighbors]
                }
            else:
                print(f"[INFO] Aucun voisin ou information récupérée pour {ip}.")
        else:
            print(f"[ERREUR] {ip} ne répond pas au ping.")
    
    # Création du tableau de la topologie (version non commentée)
    topology = []
    for device, info in network_map.items():
        for neighbor in info['neighbors']:
            topology.append([device, info['ip'], neighbor[0], neighbor[1], neighbor[2], info['protocol']])
    
    # Affichage des résultats si des informations ont été trouvées
    if topology:
        print("\nTopologie du réseau:")
        print(tabulate(topology, headers=["Équipement", "IP", "Voisin", "Interface Locale", "Interface Distante", "Protocole"], tablefmt="simple"))
        
        # Affichage du résumé des connexions (version non commentée)
        print("\nRésumé des connexions:")
        connections = defaultdict(set)
        for row in topology:
            device = row[0]
            neighbor = row[2]
            connections[device].add(neighbor)
        
        for device, neighbors in connections.items():
            print(f"{device} est connecté à: {', '.join(neighbors)}")
            
        # Création du fichier Excel (version non commentée)
        wb = Workbook()
        ws = wb.active
        ws.title = "Topologie"
        
        # Ajout des en-têtes
        headers = ["Équipement", "IP", "Voisin", "Interface Locale", "Interface Distante", "Protocole"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Ajout des données
        for row_idx, row_data in enumerate(topology, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustement de la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarde du fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"excel/topology_v1_{timestamp}.xlsx"
        wb.save(filename)
        print(f"\nTopologie exportée dans le fichier : {filename}")
        
        # Création du graphique avec Graphviz
        dot = graphviz.Graph(comment='Topologie Réseau', engine='dot')
        
        # Configuration globale du graphe
        dot.attr(overlap='false', splines='true', rankdir='TB')
        
        # Style par défaut pour les nœuds
        dot.attr('node', shape='box', style='filled', fillcolor='lightblue',
                fontname='Helvetica', fontsize='10')
        
        # Style par défaut pour les arêtes
        dot.attr('edge', fontsize='8', fontname='Helvetica')

        # Identifier automatiquement les switches de distribution et d'accès
        distribution_switches = set()
        access_switches = set()
        connections_count = defaultdict(int)

        # Compter le nombre de connexions pour chaque équipement
        for row in topology:
            device = row[0]
            neighbor = row[2]
            connections_count[device] += 1
            connections_count[neighbor] += 1

        # Classifier les switches basé sur leur nombre de connexions
        for device, count in connections_count.items():
            if count > 1:  # Si un switch a plusieurs connexions, c'est probablement un switch de distribution
                distribution_switches.add(device)
            else:
                access_switches.add(device)

        # Créer des sous-graphes pour organiser la hiérarchie
        with dot.subgraph(name='cluster_0') as dist:
            dist.attr(rank='min')  # Force les switches de distribution en haut
            for switch in distribution_switches:
                dist.node(switch, switch)

        with dot.subgraph(name='cluster_1') as acc:
            acc.attr(rank='max')  # Force les switches d'accès en bas
            for switch in access_switches:
                acc.node(switch, switch)
        
        # Ajout des connexions avec les détails des interfaces
        seen_edges = set()  # Pour éviter les doublons
        for row in topology:
            device = row[0]
            neighbor = row[2]
            local_intf = row[3]
            remote_intf = row[4]
            
            # Création d'une clé unique pour chaque liaison
            edge_key = tuple(sorted([device, neighbor]))
            if edge_key not in seen_edges:
                # Création de deux labels, un pour chaque extrémité
                dot.edge(device, neighbor, headlabel=remote_intf, taillabel=local_intf, 
                        color='darkblue', fontsize='8', labeldistance='2.0', 
                        labelfloat='false', labelangle='45')
                seen_edges.add(edge_key)
        
        # Configuration supplémentaire pour améliorer le rendu
        dot.attr(dpi='300')  # Haute résolution
        
        # Sauvegarde du graphique
        graph_filename = f"excel/topology_v1_{timestamp}"
        dot.render(graph_filename, format='png', cleanup=True)
        print(f"Représentation graphique exportée dans le fichier : {graph_filename}.png")
    else:
        print("Aucune information de topologie trouvée.")

# Exécution de la cartographie du réseau
create_network_map()
